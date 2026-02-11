#--------------------------------------------------------------
# Operations for Computing/Calculation
# Programmer: Shanqin Jin
# Email: sjin@mun.ca
# Date: 2025-11-30
#--------------------------------------------------------------

import logging
import time
from PySide6.QtCore import QObject, QThread, Signal, QMutex, QWaitCondition
from PySide6.QtWidgets import QMessageBox, QStyle
from PySide6.QtGui import QIcon

from Savitsky_Method.Savitsky_Calculation import Savitsky_Calm_Water
from SaMPH_Utils.Utils import utils

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

#==============================================================
class CalculationWorker(QThread):
    """
    Worker thread for running Savitsky calculations in the background.
    Supports pause and resume functionality.
    """
    progress_updated = Signal(dict)       # Emits result for a single speed
    calculation_finished = Signal()       # Emits when all speeds are done
    calculation_error = Signal(str)       # Emits on error
    status_message = Signal(str)          # Emits status messages for logging
    result_updated = Signal(str, float, float)  # Emits (result_type, Fn, value) for real-time updates

    def __init__(self, params, speeds):
        super().__init__()
        self.params = params
        self.speeds = speeds
        self.is_paused = False
        self.is_stopped = False
        self.mutex = QMutex()
        self.condition = QWaitCondition()
        self.current_index = 0

    def run(self):
        try:
            solver = Savitsky_Calm_Water(self.params)
            total_speeds = len(self.speeds)

            # Continue from where we left off (useful if we implement stop/resume later, 
            # but for now we just iterate)
            # If we want to support resuming from a specific index after a full stop, 
            # we'd need to pass that index in. 
            # Here we just iterate through the list.
            
            # Note: If the thread is restarted, current_index resets unless we handle it.
            # For this implementation, "Pause" just blocks the thread. "Stop" kills it.
            
            for i in range(self.current_index, total_speeds):
                self.mutex.lock()
                if self.is_stopped:
                    self.mutex.unlock()
                    break
                
                while self.is_paused:
                    self.condition.wait(self.mutex)
                    if self.is_stopped:
                        self.mutex.unlock()
                        return
                self.mutex.unlock()

                # Perform calculation
                velocity = self.speeds[i]
                result = solver.calculate_single_speed(velocity)
                
                if result:
                    result['velocity'] = velocity
                    self.progress_updated.emit(result)
                    
                    # Emit individual result updates for real-time page updates
                    fn = result.get('Fn', 0)
                    self.result_updated.emit("Rw", fn, result.get('R_hydro', 0))
                    self.result_updated.emit("Rs", fn, result.get('Rs', 0))
                    self.result_updated.emit("Ra", fn, result.get('Ra', 0))
                    self.result_updated.emit("Rt", fn, result.get('Rt', 0))
                    self.result_updated.emit("Trim", fn, result.get('trim_deg', 0))
                    self.result_updated.emit("Sinkage", fn, result.get('sinkage', 0))
                
                self.current_index = i + 1
                
                # Simulate a small delay to make the pause visible/testable for very fast calcs
                time.sleep(0.1) 

            if not self.is_stopped:
                self.calculation_finished.emit()

        except Exception as e:
            self.calculation_error.emit(str(e))

    def pause(self):
        self.mutex.lock()
        self.is_paused = True
        self.mutex.unlock()
        self.status_message.emit(f"Calculation paused at speed index {self.current_index}.")

    def resume(self):
        self.mutex.lock()
        self.is_paused = False
        self.condition.wakeAll()
        self.mutex.unlock()
        self.status_message.emit("Calculation resumed.")

    def stop(self):
        self.mutex.lock()
        self.is_stopped = True
        self.condition.wakeAll() # Wake up if paused so it can exit
        self.mutex.unlock()
        self.status_message.emit("Calculation stopped.")


#==============================================================
class Computing_Operations(QObject):
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.worker = None
        self.is_calculating = False # True if worker is alive (running or paused)
        self.is_paused = False

    # ----------------------------------------------------------------
    def handle_calculate_request(self, checked):
        """
        Handle the toolbar button click.
        checked: The state of the button (True = Pressed/Active).
        
        Logic:
        - If NOT calculating: Start new calculation.
        - If calculating and NOT paused: Pause.
        - If calculating AND paused: Resume.
        """
        
        # Scenario 1: Start new calculation
        if not self.is_calculating:
            if checked:
                self.start_calculation()
            else:
                # Should not happen normally if button was unchecked, but handle just in case
                pass
                
        # Scenario 2: Pause or Resume
        else:
            if self.is_paused:
                self.resume_calculation()
            else:
                self.pause_calculation()

    # ----------------------------------------------------------------
    def start_calculation(self):
        # 1. Validate input parameters first
        if hasattr(self.main_window, 'operations_input_page'):
            is_valid, error_message = self.main_window.operations_input_page.validate_input_parameters()
            
            if not is_valid:
                QMessageBox.critical(
                    self.main_window,
                    "Input Validation Failed",
                    error_message
                )
                self.log_message(f"Calculation cancelled: {error_message.split(chr(10))[0]}")
                self.reset_ui_state()
                return
        
        # 2. Gather Data
        input_page = getattr(self.main_window, 'page_input', None)
        if not input_page:
            QMessageBox.warning(self.main_window, "Warning", "No Input page is currently open.")
            self.reset_ui_state()
            return

        try:
            params = self.gather_parameters(input_page)
            speeds = self.gather_speeds(input_page)
            
            # Pass hull params to result page for AI evaluation
            if hasattr(self.main_window, 'operations_result_page'):
                self.main_window.operations_result_page.set_hull_params(params)
            
            if not speeds:
                QMessageBox.warning(self.main_window, "Warning", "No valid speeds defined.")
                self.reset_ui_state()
                return

            # 3. Initialize Worker
            self.results = [] # Clear previous results
            self.worker = CalculationWorker(params, speeds)
            self.worker.progress_updated.connect(self.on_progress_updated)
            self.worker.calculation_finished.connect(self.on_calculation_finished)
            self.worker.calculation_error.connect(self.on_calculation_error)
            self.worker.status_message.connect(self.log_message)
            
            # Connect result_updated signal if main window has result operations
            if hasattr(self.main_window, 'operations_result_page'):
                self.worker.result_updated.connect(
                    self.main_window.operations_result_page.handle_result_update
                )
                
                # Determine mode based on input
                mode = "scatter" if input_page.radio_discrete.isChecked() else "continuous"
                self.main_window.operations_result_page.set_mode(mode)
                
                # Clear previous results and open default pages
                self.main_window.operations_result_page.clear_all_results()
                self.main_window.operations_result_page.open_default_pages()
            
            # 4. Update UI
            self.is_calculating = True
            self.is_paused = False
            self.update_toolbar_icon(state="pause") # Show Pause icon (user can click to pause)
            
            # Log header
            self.log_message(f"Starting calculation for {len(speeds)} speeds...")
            header = f"{'V(m/s)':<10} {'Trim(deg)':<10} {'Rt(N)':<10} {'Sinkage(m)':<10} {'Lambda':<10}"
            self.log_message(header)
            self.log_message("-" * 60)
            
            # 5. Start
            self.worker.start()

        except Exception as e:
            QMessageBox.critical(self.main_window, "Error", f"Failed to start calculation: {str(e)}")
            self.reset_ui_state()

    # ----------------------------------------------------------------
    def pause_calculation(self):
        if self.worker:
            self.worker.pause()
            self.is_paused = True
            self.update_toolbar_icon(state="resume") # Show Play icon (user can click to resume)

    # ----------------------------------------------------------------
    def resume_calculation(self):
        if self.worker:
            self.worker.resume()
            self.is_paused = False
            self.update_toolbar_icon(state="pause") # Show Pause icon (user can click to pause)

    # ----------------------------------------------------------------
    def stop_calculation(self):
        if self.worker:
            self.worker.stop()
            self.worker.wait()
            self.worker = None
        self.reset_ui_state()

    # ----------------------------------------------------------------
    def on_progress_updated(self, res):
        self.results.append(res) # Store result
        line = f"{res['velocity']:<10.4f} {res['trim_deg']:<10.4f} {res['Rt']:<10.4f} {res['sinkage']:<10.4f} {res['lambda']:<10.4f}"
        self.log_message(line)

    def on_calculation_finished(self):
        self.log_message("Calculation completed successfully.")
        
        # Save results to Excel
        try:
            self.save_results_to_excel()
        except Exception as e:
            self.log_message(f"Failed to save results: {str(e)}")
        
        # Save wake profiles
        try:
            self.save_wake_profiles()
        except Exception as e:
            self.log_message(f"Failed to save wake profiles: {str(e)}")
            
        self.reset_ui_state()
        self.worker = None

    def save_results_to_excel(self):
        """
        Save the calculation results to an Excel file in the Results directory.
        """
        results_dir = utils.get_results_dir()
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"Savitsky_Results_{timestamp}.xlsx"
        filepath = results_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Calculation Results"
        
        # Headers (matching Matlab output format)
        # V  Fn  Resistance Spray-Resistance Air-Resistance Total-Resistance Trim   Sinkage   Lk   Lc   X   Y   Z   lambda   a   c   d   f  Cv
        headers = [
            "V (m/s)", "Fn", "R (N)", "Rs (N)", "Ra (N)", "Rt (N)", 
            "Trim (deg)", "Sinkage (m)", "Lk (m)", "Lc (m)", 
            "X (m)", "Y (m)", "Z (m)", "Lambda", 
            "a (m)", "c (m)", "d (m)", "f (m)", "Cv"
        ]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
        # Add data
        if self.results:
            self.log_message(f"DEBUG: Saving to Excel. First sinkage value: {self.results[0].get('sinkage', 'N/A')}")

        for res in self.results:
            row = [
                res.get('velocity', 0),         # V
                res.get('Fn', 0),               # Fn
                res.get('R_hydro', 0),          # R
                res.get('Rs', 0),               # Rs
                res.get('Ra', 0),               # Ra
                res.get('Rt', 0),               # Rt
                res.get('trim_deg', 0),         # Trim
                res.get('sinkage', 0),          # Sinkage
                res.get('Lk', 0),               # Lk
                res.get('Lc', 0),               # Lc
                res.get('X_spray', 0),          # X
                res.get('Y_spray', 0),          # Y
                res.get('Z_spray', 0),          # Z
                res.get('lambda', 0),           # lambda
                res.get('a', 0),                # a
                res.get('c', 0),                # c
                res.get('d', 0),                # d
                self.worker.params.get('f', 0), # f (from input params)
                res.get('Cv', 0)                # Cv
            ]
            ws.append(row)
            
            # Style data cells
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (ValueError, TypeError):
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filepath)
        self.log_message(f"Results saved to: {filepath}")




    def format_velocity_for_filename(self, velocity):
        """
        Format velocity for filename (e.g., 3.5666 -> 3P566).
        """
        # Convert to string with 3 decimal places, replace '.' with 'P'
        return f"{velocity:.3f}".replace('.', 'P')

    def save_wake_profiles(self):
        """
        Save wake profile .dat files for each speed in velocity-named folders.
        """
        if not self.results or not self.worker:
            return
        
        results_dir = utils.get_results_dir()
        solver = Savitsky_Calm_Water(self.worker.params)
        
        for res in self.results:
            velocity = res['velocity']
            trim_deg = res['trim_deg']
            lambda_val = res['lambda']
            Cv = res['Cv']
            
            # Calculate wake profile
            wake_data = solver.calculate_wake_profile(velocity, trim_deg, lambda_val, Cv)
            
            # Create velocity-named folder
            vel_str = self.format_velocity_for_filename(velocity)
            vel_folder = results_dir / vel_str
            vel_folder.mkdir(parents=True, exist_ok=True)
            
            # Create .dat filename
            dat_filename = f"{vel_str}_WakeProfile.dat"
            dat_filepath = vel_folder / dat_filename
            
            # Write .dat file
            with open(dat_filepath, 'w') as f:
                # Header
                f.write("# X/B  Centerline_Wake_Profile/B  Quarterbeam_Wake_Profile/B\n")
                
                # Data rows
                for i in range(len(wake_data['X'])):
                    f.write(f"{wake_data['X'][i]:.6f} {wake_data['Centerline_H'][i]:.6f} {wake_data['Quarterbeam_H'][i]:.6f}\n")
        
        self.log_message(f"Wake profiles saved for {len(self.results)} velocities.")

    def on_calculation_error(self, error_msg):
        self.log_message(f"Error: {error_msg}")
        QMessageBox.critical(self.main_window, "Calculation Error", error_msg)
        self.reset_ui_state()
        self.worker = None

    # ----------------------------------------------------------------
    def reset_ui_state(self):
        self.is_calculating = False
        self.is_paused = False
        # Reset toolbar button to initial state (Play icon, unchecked)
        # We need to access the action directly or via signal
        # Since we are handling the signal, we might need to manually update the action state
        # if the signal logic in ToolBar doesn't cover external resets.
        
        # Accessing toolbar action directly for now
        if hasattr(self.main_window, 'tool_bar'):
            action = self.main_window.tool_bar.action_calculate
            # Block signals to prevent triggering handle_calculate_request again
            action.blockSignals(True)
            action.setChecked(False)
            action.setText("Calculate")
            action.setToolTip("Calculate")
            action.setIcon(QIcon(utils.local_resource_path("SaMPH_Images/WIN11-Icons/icons8-play-100.png")))
            action.blockSignals(False)

    def update_toolbar_icon(self, state):
        if hasattr(self.main_window, 'tool_bar'):
            action = self.main_window.tool_bar.action_calculate
            if state == "pause":
                # User is running, show Pause button
                action.setText("Pause")
                action.setToolTip("Pause Calculation")
                action.setIcon(QIcon(utils.local_resource_path("SaMPH_Images/WIN11-Icons/icons8-stop-100.png")))
            elif state == "resume":
                # User is paused, show Resume (Play) button
                action.setText("Resume")
                action.setToolTip("Resume Calculation")
                action.setIcon(QIcon(utils.local_resource_path("SaMPH_Images/WIN11-Icons/icons8-play-100.png")))

    def log_message(self, msg):
        if hasattr(self.main_window, 'log_window'):
            self.main_window.log_window.log_message(msg)

    # ----------------------------------------------------------------
    def handle_clear_request(self):
        """
        Handle clear button click.
        Clear all inputs, results, and logs after user confirmation.
        """
        # Show confirmation dialog
        reply = QMessageBox.question(
            self.main_window,
            "Clear All Data",
            "Are you sure you want to clear all inputs, results, and logs?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No  # Default to "No" for safety
        )
        
        # Only proceed if user clicked "Yes"
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # 1. Clear inputs
        input_page = getattr(self.main_window, 'page_input', None)
        if input_page:
            input_page.reset_parameters()
        
        # 2. Clear results
        self.main_window.operations_result_page.clear_all_results()
        
        # 3. Clear log
        log_window = getattr(self.main_window, 'log_window', None)
        if log_window:
            log_window.clear_log()
            
        # 4. Reset UI state
        self.reset_ui_state()
        self.log_message("All data cleared.")

    def gather_parameters(self, input_page):
        params = {}
        # Constants
        params['g'] = float(input_page.inputs['acceleration_of_gravity'].text())
        params['rho'] = float(input_page.inputs['density_of_water'].text())
        params['nu'] = float(input_page.inputs['kinematic_viscosity_of_water'].text())
        
        # Hull Parameters
        params['ship_length'] = float(input_page.inputs['ship_length'].text())
        params['ship_beam'] = float(input_page.inputs['ship_beam'].text())
        
        # User input is Displacement (Weight), so Mass = Displacement / g
        displacement_val = float(input_page.inputs['displacement'].text())
        params['mass'] = displacement_val / params['g']
        
        params['beta'] = float(input_page.inputs['deadrise_angle'].text())
        params['lcg'] = float(input_page.inputs['longitudinal_center_of_gravity'].text())
        params['vcg'] = float(input_page.inputs['vertical_center_of_gravity'].text())
        params['draft'] = float(input_page.inputs['mean_draft'].text())
        params['frontal_area'] = float(input_page.inputs['frontal_area_of_ship'].text())
        
        # Optional/Fixed parameters
        params['f'] = 0.0
        params['epsilon'] = 0.0
        return params

    def gather_speeds(self, input_page):
        speeds = []
        if input_page.radio_discrete.isChecked():
            speed_text = input_page.speed_input.discrete_speeds.text()
            speeds = [float(s.strip()) for s in speed_text.split(',') if s.strip()]
        else:
            v_start = float(input_page.speed_input.continuous_initial.text())
            v_end = float(input_page.speed_input.continuous_final.text())
            v_step = float(input_page.speed_input.continuous_increment.text())
            import numpy as np
            if v_step > 0:
                speeds = np.arange(v_start, v_end + v_step*0.001, v_step).tolist()
            else:
                speeds = [v_start]
        return speeds
