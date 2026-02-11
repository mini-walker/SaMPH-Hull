#--------------------------------------------------------------
# This file contains operations for the main window
# Programmer: Shanqin Jin
# Email: sjin@mun.ca
# Date: 2025-11-27  
#--------------------------------------------------------------

import logging
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QMessageBox

# Excel file handling
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font

#==============================================================
class InputPage_Operations:

    """
    This class contains all the operation methods for the input page.
    It handles material change.
    """
    
    def __init__(self, main_window):

        """Initialize with reference to the main window"""

        self.main_window = main_window
    
    # ================================================================
    # Parameter Validation Function
    # ================================================================
    
    def validate_input_parameters(self):
        """
        Validate all input parameters before calculation.
        
        Returns:
            tuple: (is_valid: bool, error_message: str)
            - is_valid: True if all parameters are valid, False otherwise
            - error_message: Detailed error message if validation fails, empty string if valid
        """
        
        # Get the active input page
        input_page = getattr(self.main_window, 'page_input', None)
        
        if not input_page:
            return False, "Input page not found. Please open the Input page first."
        
        errors = []
        
        # ============ Section 1: Constants Validation ============
        constants_params = [
            ('acceleration_of_gravity', 'Acceleration of gravity', (0, 1e18)),
            ('density_of_water', 'Density of water', (0, 1e18)),
            ('kinematic_viscosity_of_water', 'Kinematic viscosity of water', (1e-18, 1e18))
        ]
        
        for param_key, param_name, (min_val, max_val) in constants_params:
            if param_key not in input_page.inputs:
                errors.append(f"{param_name}: Field not found in form")
                continue
            
            input_widget = input_page.inputs[param_key]
            value_text = input_widget.text().strip()
            
            # Check if empty
            if not value_text:
                errors.append(f"{param_name}: Field is empty")
                continue
            
            # Check if it's a valid float
            try:
                value = float(value_text)
            except ValueError:
                errors.append(f"{param_name}: '{value_text}' is not a valid number (expected float)")
                continue
            
            # Check if within reasonable range
            if value <= min_val or value >= max_val:
                errors.append(f"{param_name}: Value {value} is out of acceptable range [{min_val}, {max_val}]")
            
            # Check for negative values where not allowed
            if param_key in ['density_of_water', 'kinematic_viscosity_of_water'] and value < 0:
                errors.append(f"{param_name}: Cannot be negative (got {value})")
        
        # ============ Section 2: Hull Parameters Validation ============
        hull_params = [
            ('ship_length', 'Ship length', (0.00001, 1e18)),
            ('ship_beam', 'Ship beam', (1e-8, 1e18)),
            ('mean_draft', 'Mean draft', (1e-8, 1000)),
            ('displacement', 'Displacement', (1e-8, 1e18)),
            ('deadrise_angle', 'Deadrise angle', (-90, 90)),
            ('frontal_area_of_ship', 'Frontal area of ship', (0.0000001, 1e18)),
            ('longitudinal_center_of_gravity', 'Longitudinal center of gravity', (-1e18, 1e18)),
            ('vertical_center_of_gravity', 'Vertical center of gravity', (-1e18, 1e18))
        ]
        
        for param_key, param_name, (min_val, max_val) in hull_params:
            if param_key not in input_page.inputs:
                errors.append(f"{param_name}: Field not found in form")
                continue
            
            input_widget = input_page.inputs[param_key]
            value_text = input_widget.text().strip()
            
            # Check if empty
            if not value_text:
                errors.append(f"{param_name}: Field is empty")
                continue
            
            # Check if it's a valid float
            try:
                value = float(value_text)
            except ValueError:
                errors.append(f"{param_name}: '{value_text}' is not a valid number (expected float)")
                continue
            
            # Check if within reasonable range
            if value <= min_val or value >= max_val:
                errors.append(f"{param_name}: Value {value} is out of acceptable range ({min_val}, {max_val})")
            
            # Special checks for specific parameters
            if param_key == 'deadrise_angle' and not (-89 < value < 89):
                errors.append(f"{param_name}: Should be between -89° and 89° (got {value}°)")
        
        # ============ Section 3: Speed Configuration Validation ============
        try:
            speeds = input_page.speed_input.get_speed_values()
            
            if not speeds:
                errors.append("Speed Configuration: No valid speeds defined")
            elif len(speeds) == 0:
                errors.append("Speed Configuration: Speed list is empty")
            else:
                # Check speed values are within reasonable range
                for speed in speeds:
                    if speed < 0 or speed > 100:
                        errors.append(f"Speed Configuration: Speed value {speed} m/s is out of range [0, 100]")
                        break
        
        except ValueError as e:
            errors.append(f"Speed Configuration: {str(e)}")
        
        # ============ Return Results ============
        if errors:
            # Combine all errors into a single message
            error_message = "Parameter Validation Failed:\n\n" + "\n".join(errors)
            return False, error_message
        else:
            return True, ""
    
    # ================================================================
    # ----------------------------------------------------------------
    # The function to handle material library
    def handle_material_change(self, text: str):

        # Send message to log window
        # Note: using log_window instead of log_widget as per GUI_SaMPH.py
        if text == "": 
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Please enter the material properties manually.")
        else:
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Change material to: {text}")

        # Get the active input page
        # The Operation_MainWindow sets self.main_window.page_input when creating it
        input_page = getattr(self.main_window, 'page_input', None)
        
        if not input_page:
            self.main_window.log_window.log_message(f"Input page not found.", level=logging.WARNING)
            return
            
        if "Fresh water" in text or "淡水" in text: 
            if "density_of_water" in input_page.inputs:
                input_page.inputs["density_of_water"].setText("998.2")
            if "kinematic_viscosity_of_water" in input_page.inputs:
                input_page.inputs["kinematic_viscosity_of_water"].setText("0.00000100")

        elif "Sea water" in text or "海水" in text: 
            if "density_of_water" in input_page.inputs:
                input_page.inputs["density_of_water"].setText("1025.0")
            if "kinematic_viscosity_of_water" in input_page.inputs:
                input_page.inputs["kinematic_viscosity_of_water"].setText("0.000001050")
    # ----------------------------------------------------------------

    # ----------------------------------------------------------------
    def save_input_data_to_csv(self):
        """Save Input Page data to Excel file with formatting"""
        
        # Get the active input page
        input_page = getattr(self.main_window, 'page_input', None)
        
        if not input_page:
            QMessageBox.warning(
                self.main_window,
                "Warning",
                "No Input page is currently open. Please open an Input page first."
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message("Save failed: Input page not found", level=logging.WARNING)
            return
        
        # Open file dialog to select save location
        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window,
            "Save Input Data",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Input Data"
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header font
            header_font = Font(bold=True)
            
            # Write header
            headers = ['Section', 'Parameter', 'Value']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            row_idx = 2
            
            # Helper function to format parameter name: replace _ with space
            def format_param_name(param_name):
                return param_name.replace('_', ' ').title()
            
            # Section 1: Constants
            constants_params = [
                'acceleration_of_gravity',
                'density_of_water',
                'kinematic_viscosity_of_water'
            ]
            
            for param in constants_params:
                if param in input_page.inputs:
                    value = input_page.inputs[param].text().strip()
                    ws.cell(row=row_idx, column=1, value='Constants').border = thin_border
                    ws.cell(row=row_idx, column=2, value=format_param_name(param)).border = thin_border
                    ws.cell(row=row_idx, column=3, value=value).border = thin_border
                    row_idx += 1
            
            # Also save material combo selection
            material_value = input_page.material_combo.currentText()
            ws.cell(row=row_idx, column=1, value='Constants').border = thin_border
            ws.cell(row=row_idx, column=2, value='Material Preset').border = thin_border
            ws.cell(row=row_idx, column=3, value=material_value).border = thin_border
            row_idx += 1
            
            # Section 2: Speed Configuration
            speed_mode = 'discrete' if input_page.radio_discrete.isChecked() else 'continuous'
            ws.cell(row=row_idx, column=1, value='Speed Configuration').border = thin_border
            ws.cell(row=row_idx, column=2, value='Mode').border = thin_border
            ws.cell(row=row_idx, column=3, value=speed_mode).border = thin_border
            row_idx += 1
            
            # Save discrete mode parameters
            discrete_speeds = input_page.speed_input.discrete_speeds.text().strip()
            ws.cell(row=row_idx, column=1, value='Speed Configuration').border = thin_border
            ws.cell(row=row_idx, column=2, value='Discrete Speeds').border = thin_border
            ws.cell(row=row_idx, column=3, value=discrete_speeds).border = thin_border
            row_idx += 1
            
            # Save continuous mode parameters
            continuous_initial = input_page.speed_input.continuous_initial.text().strip()
            continuous_final = input_page.speed_input.continuous_final.text().strip()
            continuous_increment = input_page.speed_input.continuous_increment.text().strip()
            
            ws.cell(row=row_idx, column=1, value='Speed Configuration').border = thin_border
            ws.cell(row=row_idx, column=2, value='Continuous Initial').border = thin_border
            ws.cell(row=row_idx, column=3, value=continuous_initial).border = thin_border
            row_idx += 1
            
            ws.cell(row=row_idx, column=1, value='Speed Configuration').border = thin_border
            ws.cell(row=row_idx, column=2, value='Continuous Final').border = thin_border
            ws.cell(row=row_idx, column=3, value=continuous_final).border = thin_border
            row_idx += 1
            
            ws.cell(row=row_idx, column=1, value='Speed Configuration').border = thin_border
            ws.cell(row=row_idx, column=2, value='Continuous Increment').border = thin_border
            ws.cell(row=row_idx, column=3, value=continuous_increment).border = thin_border
            row_idx += 1
            
            # Section 3: Hull Parameters
            hull_params = [
                'ship_length',
                'ship_beam',
                'mean_draft',
                'displacement',
                'deadrise_angle',
                'frontal_area_of_ship',
                'longitudinal_center_of_gravity',
                'vertical_center_of_gravity'
            ]
            
            for param in hull_params:
                if param in input_page.inputs:
                    value = input_page.inputs[param].text().strip()
                    ws.cell(row=row_idx, column=1, value='Hull Parameters').border = thin_border
                    ws.cell(row=row_idx, column=2, value=format_param_name(param)).border = thin_border
                    ws.cell(row=row_idx, column=3, value=value).border = thin_border
                    row_idx += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Add padding, max 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            QMessageBox.information(
                self.main_window,
                "Success",
                f"Input data saved successfully to:\n{file_path}"
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Input data saved to: {file_path}")
                
        except Exception as e:
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Failed to save input data:\n{str(e)}"
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Save failed: {str(e)}", level=logging.ERROR)
    # ----------------------------------------------------------------

    # ----------------------------------------------------------------
    def load_input_data_from_csv(self):
        """Load Input Page data from Excel file"""
        
        # Get the active input page
        input_page = getattr(self.main_window, 'page_input', None)
        
        if not input_page:
            QMessageBox.warning(
                self.main_window,
                "Warning",
                "No Input page is currently open. Please open an Input page first."
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message("Load failed: Input page not found", level=logging.WARNING)
            return
        
        # Open file dialog to select Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self.main_window,
            "Load Input Data",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            # Load workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Helper function to convert parameter name: replace spaces with _
            def parse_param_name(display_name):
                return display_name.lower().replace(' ', '_')
            
            # Read and populate data
            material_preset = None
            speed_mode = None
            
            # Skip header row, start from row 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3:
                    continue  # Skip empty/invalid rows
                
                section, parameter, value = row[0], row[1], row[2]
                
                # Skip rows with None values
                if not section or not parameter:
                    continue
                
                # Convert value to string
                value = str(value) if value is not None else ""
                
                # Handle Constants section
                if section == 'Constants':
                    if parameter == 'Material Preset':
                        material_preset = value
                    else:
                        # Convert parameter name back to variable format
                        param_key = parse_param_name(parameter)
                        if param_key in input_page.inputs:
                            input_page.inputs[param_key].setText(value)
                
                # Handle Speed Configuration section
                elif section == 'Speed Configuration':
                    if parameter == 'Mode':
                        speed_mode = value
                    elif parameter == 'Discrete Speeds':
                        input_page.speed_input.discrete_speeds.setText(value)
                    elif parameter == 'Continuous Initial':
                        input_page.speed_input.continuous_initial.setText(value)
                    elif parameter == 'Continuous Final':
                        input_page.speed_input.continuous_final.setText(value)
                    elif parameter == 'Continuous Increment':
                        input_page.speed_input.continuous_increment.setText(value)
                
                # Handle Hull Parameters section
                elif section == 'Hull Parameters':
                    # Convert parameter name back to variable format
                    param_key = parse_param_name(parameter)
                    if param_key in input_page.inputs:
                        input_page.inputs[param_key].setText(value)
            
            # Set material combo box
            if material_preset:
                index = input_page.material_combo.findText(material_preset)
                if index >= 0:
                    input_page.material_combo.setCurrentIndex(index)
            
            # Set speed mode
            if speed_mode == 'discrete':
                input_page.radio_discrete.setChecked(True)
            elif speed_mode == 'continuous':
                input_page.radio_continuous.setChecked(True)
            
            # Update previews
            input_page.speed_input.update_discrete_preview()
            input_page.speed_input.update_continuous_preview()
            
            QMessageBox.information(
                self.main_window,
                "Success",
                f"Input data loaded successfully from:\n{file_path}"
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Input data loaded from: {file_path}")
                
        except FileNotFoundError:
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"File not found:\n{file_path}"
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Load failed: File not found", level=logging.ERROR)
        except Exception as e:
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Failed to load input data:\n{str(e)}"
            )
            if hasattr(self.main_window, 'log_window'):
                self.main_window.log_window.log_message(f"Load failed: {str(e)}", level=logging.ERROR)
    # ----------------------------------------------------------------

#==============================================================
